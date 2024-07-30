import os
import datetime
import sqlite3


class Person:
    def __init__(
        self,
        Serial_Number=None,
        NIC=None,
        Rent=None,
        Rentel_Name=None,
        Due_Date=None,
        Received_Rent=None,
        Balance_Rent=None,
        Electric_Bill=None,
        Electricity_Meter_Number=None,
        Electricity_Account_Number=None,
        Consumer_Number=None,
        Electricity_Meter_Name=None,
        Gas_Costumer_Number=None,
        Gas_Meter_Number=None,
        Advance_Amount=None,
        Building=None,
        Gas_Bill=None,
    ):
        self.current_month, self.previous_month = self.get_current_and_previous_month()
        self.setup_directories()
        self.setup_database()

        if all(v is not None for v in locals().values()):
            today = datetime.date.today().strftime("%Y-%m-%d")
            data = (
                Serial_Number,
                NIC,
                Rent,
                Rentel_Name,
                Due_Date,
                Received_Rent,
                Balance_Rent,
                Electric_Bill,
                Electricity_Meter_Number,
                Electricity_Account_Number,
                Consumer_Number,
                Electricity_Meter_Name,
                Gas_Costumer_Number,
                Gas_Meter_Number,
                Advance_Amount,
                Building,
                Gas_Bill,
                today,
            )
            self.insert_data(data)

    def get_current_and_previous_month(self):
        today = datetime.date.today()
        current_month = today.strftime("%B_%Y")
        first_day_of_current_month = today.replace(day=1)
        last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(
            days=1
        )
        previous_month = last_day_of_previous_month.strftime("%B_%Y")
        return current_month, previous_month

    def setup_directories(self):
        os.makedirs(f"data/{self.current_month}", exist_ok=True)
        previous_month_dir = f"data/{self.previous_month}"
        if not os.path.exists(previous_month_dir):
            os.makedirs(previous_month_dir)

    def setup_database(self):
        self.conn = sqlite3.connect(f"data/{self.current_month}/persondata.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS persondata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                Serial_Number TEXT,
                NIC TEXT,
                Rent REAL,
                Rentel_Name TEXT,
                Due_Date TEXT,
                Received_Rent REAL,
                Balance_Rent REAL,
                Electric_Bill REAL,
                Electricity_Meter_Number TEXT,
                Electricity_Account_Number TEXT,
                Consumer_Number TEXT,
                Electricity_Meter_Name TEXT,
                Gas_Costumer_Number TEXT,
                Gas_Meter_Number TEXT,
                Advance_Amount REAL,
                Building TEXT,
                Gas_Bill REAL,
                Date_Added TEXT
            )
        """
        )
        self.conn.commit()

    def insert_data(self, data):
        self.cursor.execute(
            """
            INSERT INTO persondata (
                Serial_Number,
                NIC,
                Rent,
                Rentel_Name,
                Due_Date,
                Received_Rent,
                Balance_Rent,
                Electric_Bill,
                Electricity_Meter_Number,
                Electricity_Account_Number,
                Consumer_Number,
                Electricity_Meter_Name,
                Gas_Costumer_Number,
                Gas_Meter_Number,
                Advance_Amount,
                Building,
                Gas_Bill,
                Date_Added
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            data,
        )
        self.conn.commit()

    def get_all_persons(self):
        self.cursor.execute("SELECT * FROM persondata")
        rows = self.cursor.fetchall()
        return rows

    def delete_person(self, rentel_name):
        self.cursor.execute(
            "DELETE FROM persondata WHERE Rentel_Name = ?", (rentel_name,)
        )
        self.conn.commit()

    def update_person(self, rentel_name, **kwargs):
        columns = ", ".join(f"{key} = ?" for key in kwargs.keys())
        values = list(kwargs.values()) + [rentel_name]
        self.cursor.execute(
            f"UPDATE persondata SET {columns} WHERE Rentel_Name = ?", values
        )
        self.conn.commit()

    def __del__(self):
        self.conn.close()


class PersonManager:
    def __init__(self):
        self.person = Person()

    def ask_method(self):
        while True:
            action = (
                input(
                    "What would you like to do? (add, delete, update, get_all, exit): "
                )
                .strip()
                .lower()
            )
            if action == "add":
                self.add_person()
            elif action == "delete":
                self.delete_person()
            elif action == "update":
                self.update_person()
            elif action == "get_all":
                self.get_all_persons()
            elif action == "exit":
                break
            else:
                print("Invalid option. Please try again.")

    def add_person(self):
        data = {}
        data["Serial_Number"] = input("Enter Serial Number: ")
        data["NIC"] = input("Enter NIC: ")
        data["Rent"] = float(input("Enter Rent: "))
        data["Rentel_Name"] = input("Enter Rentel Name: ")
        data["Due_Date"] = input("Enter Due Date (YYYY-MM-DD): ")
        data["Received_Rent"] = float(input("Enter Received Rent: "))
        data["Balance_Rent"] = float(input("Enter Balance Rent: "))
        data["Electric_Bill"] = float(input("Enter Electric Bill: "))
        data["Electricity_Meter_Number"] = input("Enter Electricity Meter Number: ")
        data["Electricity_Account_Number"] = input("Enter Electricity Account Number: ")
        data["Consumer_Number"] = input("Enter Consumer Number: ")
        data["Electricity_Meter_Name"] = input("Enter Electricity Meter Name: ")
        data["Gas_Costumer_Number"] = input("Enter Gas Costumer Number: ")
        data["Gas_Meter_Number"] = input("Enter Gas Meter Number: ")
        data["Advance_Amount"] = float(input("Enter Advance Amount: "))
        data["Building"] = input("Enter Building: ")
        data["Gas_Bill"] = float(input("Enter Gas Bill: "))

        person = Person(**data)
        print("Person added successfully!")

    def delete_person(self):
        rentel_name = input("Enter Rentel Name to delete: ")
        self.person.delete_person(rentel_name)
        print(f"Person with Rentel Name '{rentel_name}' deleted successfully!")

    def update_person(self):
        rentel_name = input("Enter Rentel Name to update: ")
        updates = {}
        while True:
            field = (
                input("Enter field to update (or 'done' to finish): ").strip().lower()
            )
            if field == "done":
                break
            value = input(f"Enter new value for {field}: ")
            updates[field] = value
        self.person.update_person(rentel_name, **updates)
        print(f"Person with Rentel Name '{rentel_name}' updated successfully!")

    def get_all_persons(self):
        persons = self.person.get_all_persons()
        for person in persons:
            print(person)


if __name__ == "__main__":
    manager = PersonManager()
    manager.ask_method()

exit()
# Summary is a class that use other methods to calculate the total rent, balance and recieving amounts of given json file.
import json


class Summary:
    def __init__(self, filename):
        self.filename = filename
        self.json: dict = self.read_json()
        # print(self.read_json())
        self.total_rent = self.calculate_total_rent()

    def read_json(self) -> dict:
        with open(self.filename, "r") as file:
            data = json.load(file)
        return data

    def calculate_total_rent(self) -> int:
        total_rent = 0
        total_balance = 0
        total_recieving_amounts = 0
        for person, person_data in self.json.items():
            for key, value in person_data.items():
                if key == "Rent":
                    print(
                        {
                            person: {
                                "Rent": person_data["Rent"],
                                "Balance_Rent": person_data["Balance_Rent"],
                                "Received_Rent": person_data["Received_Rent"],
                            }
                        }
                    )
                    total_rent += int(value)
                    # print(int(value))
                elif key == "Balance_Rent":
                    total_balance += int(value)

                elif key == "Received_Rent":
                    total_recieving_amounts += int(value)
                    # total_rent -= int(value)

        total_rent -= int(total_recieving_amounts)
        return {
            "Total_Rent": total_rent,
            "Balance Rent": total_balance,
            "Received_Rent": total_recieving_amounts,
        }


if __name__ == "__main__":
    summary = Summary(
        r"C:\Users\Black\OneDrive\Desktop\Rent Recorder\data\July_2024\persondata.json"
    )
    print(f"Total rent: {summary.total_rent}")
    # Add more methods to calculate balance and recieving amounts here.
    # For example, balance = total_rent - total_recieving_amounts, etc.
    # print(f"Balance: {summary.calculate_balance()}")
    # print(f"Total recieving amounts: {summary.calculate_total_recieving_amounts()}")
    # etc.

exit()
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QPushButton,
)
import sys
import openpyxl

path = "Rent sheet  JUNE(AutoRecovered).xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active


def replace_at_index(tup, index, value):
    return tup[:index] + (value,) + tup[index + 1 :]


class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel data to QTableWidget")

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)
        add_person = QPushButton("Add Person")
        layout.addWidget(add_person)

        add_person.clicked.connect(self.add_person)

        self.load_data(sheet)

    def add_person(self):
        # global sheet, workbook, path
        sheet.append(["person", "name"])
        workbook.save("person.xlsx")
        path2 = "person.xlsx"
        workbook2 = openpyxl.load_workbook(path2)
        sheet2 = workbook2.active
        self.load_data(sheet2)

    def load_data(self, sheet):

        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(
                    row_index, col_index, QTableWidgetItem(str(value))
                )
                col_index += 1
            row_index += 1


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Main()
    window.showMaximized()
    app.exec_()

# data={
# houseno,
# rent,
# rentalname,
# due,
# recivdata,
# balancerent,
# elecbill,
# gassbil,
# Nic,
# advancebill,
# gasmeterno,
# gascostumerno,
# electmetername,
# electmeterno,
# electacountno,
# consumerno,
# }

# allmonthshistry
