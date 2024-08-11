import json
import os
import sqlite3
import datetime
import pandas as pd
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import qt_material
import sys
from home import Ui_MainWindow
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from KK_Moosa_Plot_no_72 import Ui_Form

# Reset database function for initializing database tables
def reset_database():
    conn = sqlite3.connect("property_management.db")
    cursor = conn.cursor()

    cursor.execute('DROP TABLE IF EXISTS Tenants')
    cursor.execute('DROP TABLE IF EXISTS RentHistory')

    cursor.execute('''CREATE TABLE IF NOT EXISTS Tenants (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Serial_Number TEXT,
                        NIC TEXT,
                        Rent REAL,
                        Rentel_Name TEXT,
                        Due_Date TEXT,
                        Received_Rent REAL,
                        Balance_Rent REAL,
                        Electric_Bill TEXT,
                        Electricity_Meter_Number TEXT,
                        Electricity_Account_Number TEXT,
                        Consumer_Number TEXT,
                        Electricity_Meter_Name TEXT,
                        Gas_Costumer_Number TEXT,
                        Gas_Meter_Number TEXT,
                        Advance_Amount REAL,
                        Building TEXT,
                        Gas_Bill TEXT,
                        date_added TEXT)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS RentHistory (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        tenant_id INTEGER,
                        date TEXT,
                        Rent REAL,
                        Received_Rent REAL,
                        Balance_Rent REAL,
                        FOREIGN KEY (tenant_id) REFERENCES Tenants(id))''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS expenses (
                        title TEXT PRIMARY KEY, 
                        expense REAL, 
                        date_added TEXT)''')

    conn.commit()
    conn.close()

# Class for managing expenses
class Expenses:
    def __init__(self):
        self.conn = sqlite3.connect("property_management.db")
        self.cursor = self.conn.cursor()
        self.create_table()

    def create_table(self):
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS expenses 
                            (title TEXT PRIMARY KEY, 
                            expense REAL, 
                            date_added TEXT)""")
        self.conn.commit()

    def add_expense(self, Title, Expense):
        self.cursor.execute("INSERT INTO expenses VALUES (?,?,?)", 
                            (Title, Expense, datetime.datetime.now().strftime("%Y-%m-%d")))
        self.conn.commit()

    def delete_expense(self, title):
        self.cursor.execute("DELETE FROM expenses WHERE title =?", (title,))
        self.conn.commit()

    def update_expense(self, title, new_expense, new_title):
        self.delete_expense(title)
        self.add_expense(new_title, new_expense)

    def get_expense(self, title):
        self.cursor.execute("SELECT * FROM expenses WHERE title =?", (title,))
        return self.cursor.fetchone()

    def get_all_expenses(self):
        self.cursor.execute("SELECT * FROM expenses")
        return self.cursor.fetchall()

# Class for managing tenant information
class Person:
    def __init__(self, Serial_Number, NIC, Rent, Rentel_Name, Due_Date, Received_Rent, Balance_Rent, Electric_Bill,
                 Electricity_Meter_Number, Electricity_Account_Number, Consumer_Number, Electricity_Meter_Name,
                 Gas_Costumer_Number, Gas_Meter_Number, Advance_Amount, Building, Gas_Bill):
        self.conn = sqlite3.connect("property_management.db")
        self.cursor = self.conn.cursor()
        self.create_tables()

        today = datetime.date.today().strftime("%Y-%m-%d")
        self.add_person(Serial_Number, NIC, Rent, Rentel_Name, Due_Date, Received_Rent, Balance_Rent, Electric_Bill,
                        Electricity_Meter_Number, Electricity_Account_Number, Consumer_Number, Electricity_Meter_Name,
                        Gas_Costumer_Number, Gas_Meter_Number, Advance_Amount, Building, Gas_Bill, today)

    def create_tables(self):
        self.cursor.execute('''CREATE TABLE IF NOT EXISTS Tenants (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                Serial_Number TEXT,
                                NIC TEXT,
                                Rent REAL,
                                Rentel_Name TEXT,
                                Due_Date TEXT,
                                Received_Rent REAL,
                                Balance_Rent REAL,
                                Electric_Bill TEXT,
                                Electricity_Meter_Number TEXT,
                                Electricity_Account_Number TEXT,
                                Consumer_Number TEXT,
                                Electricity_Meter_Name TEXT,
                                Gas_Costumer_Number TEXT,
                                Gas_Meter_Number TEXT,
                                Advance_Amount REAL,
                                Building TEXT,
                                Gas_Bill TEXT,
                                date_added TEXT)''')

        self.cursor.execute('''CREATE TABLE IF NOT EXISTS RentHistory (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                tenant_id INTEGER,
                                date TEXT,
                                Rent REAL,
                                Received_Rent REAL,
                                Balance_Rent REAL,
                                FOREIGN KEY (tenant_id) REFERENCES Tenants(id))''')
        self.conn.commit()

    def add_person(self, Serial_Number, NIC, Rent, Rentel_Name, Due_Date, Received_Rent, Balance_Rent, Electric_Bill,
                   Electricity_Meter_Number, Electricity_Account_Number, Consumer_Number, Electricity_Meter_Name,
                   Gas_Costumer_Number, Gas_Meter_Number, Advance_Amount, Building, Gas_Bill, date_added):
        self.cursor.execute('''INSERT INTO Tenants (Serial_Number, NIC, Rent, Rentel_Name, Due_Date, Received_Rent, 
                              Balance_Rent, Electric_Bill, Electricity_Meter_Number, Electricity_Account_Number, 
                              Consumer_Number, Electricity_Meter_Name, Gas_Costumer_Number, Gas_Meter_Number, 
                              Advance_Amount, Building, Gas_Bill, date_added) 
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (Serial_Number, NIC, Rent, Rentel_Name, Due_Date, Received_Rent, Balance_Rent, Electric_Bill,
                             Electricity_Meter_Number, Electricity_Account_Number, Consumer_Number, Electricity_Meter_Name,
                             Gas_Costumer_Number, Gas_Meter_Number, Advance_Amount, Building, Gas_Bill, date_added))
        tenant_id = self.cursor.lastrowid
        self.add_rent_history(tenant_id, date_added, Rent, Received_Rent, Balance_Rent)
        self.conn.commit()

    def add_rent_history(self, tenant_id, date, Rent, Received_Rent, Balance_Rent):
        self.cursor.execute('''INSERT INTO RentHistory (tenant_id, date, Rent, Received_Rent, Balance_Rent) 
                               VALUES (?, ?, ?, ?, ?)''', (tenant_id, date, Rent, Received_Rent, Balance_Rent))
        self.conn.commit()

    def get_person_by_building(self, building):
        self.cursor.execute('SELECT * FROM Tenants WHERE Building = ?', (building,))
        return self.cursor.fetchall()

    def update_person(self, person_data):
        self.cursor.execute('''UPDATE Tenants SET NIC = ?, Rent = ?, Rentel_Name = ?, Due_Date = ?, 
                               Received_Rent = ?, Balance_Rent = ?, Electric_Bill = ?, Gas_Bill = ?, 
                               Electricity_Meter_Number = ?, Electricity_Account_Number = ?, 
                               Consumer_Number = ?, Electricity_Meter_Name = ?, Gas_Costumer_Number = ?, 
                               Gas_Meter_Number = ?, Advance_Amount = ?, Building = ? 
                               WHERE Serial_Number = ?''', person_data)
        self.conn.commit()

# Main GUI class
class XLSX(QMainWindow):
    def __init__(self):
        super().__init__()
        self.homeui = Ui_MainWindow()
        self.Expense_Sys = Expenses()
        self.homeui.setupUi(self)
        self.homeui.Tab_window.currentChanged.connect(self.taber)
        self.SetupUI()

        # UI element connections
        self.homeui.action.triggered.connect(self.add_building)
        self.homeui.action_2.triggered.connect(self.remove_building)
        self.homeui.add_exp_btn.clicked.connect(self.add_exp_method)
        self.read_buildings_file()
        self.homeui.Electric_Bill.clicked.connect(self.electric_bill_method)
        self.homeui.Gas_Bill.clicked.connect(self.gas_bill_method)
        self.homeui.buildingcombobox.currentTextChanged.connect(self.Add_Building_dialog)
        self.homeui.menuPrint.triggered.connect(lambda: os.system("print " + f"RentData.xlsx"))
        self.homeui.Add_Person_btn.clicked.connect(self.Add_Person_func)
        self.homeui.menuConvert_To_Excel.mousePressEvent = self.convert_to_excel

        # Apply theme settings
        self.setting = QSettings("Rent Recorder", "Theme")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
        self.setStyleSheet(self.styleSheet() + '*{font: 22pt "Cascadia Code";}')

        # Timer for search functionality
        self.searchtimer = QTimer(self.homeui.searchedit)
        self.searchtimer.timeout.connect(self.setupui_search)
        self.searchtimer.start(1000)
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def add_exp_method(self):
        title = self.homeui.title_lineedit.text()
        amount = self.homeui.exp_amount_lineedit.text()
        self.Expense_Sys.add_expense(title, int(amount))
        self.taber(2)

    def add_building(self):
        dialog = QDialog(self)
        add_building_line_edit = QLineEdit(dialog)
        add_building_line_edit.returnPressed.connect(lambda _="", name=add_building_line_edit: self.Add_Building(name))
        add_building_line_edit.setPlaceholderText("Enter Building Name")
        add_building_line_edit.move(10, 10)
        dialog.setWindowTitle("Add Building")
        add_building_line_edit.show()
        dialog.exec_()

    def read_buildings_file(self):
        index = self.homeui.buildingcombobox.currentIndex()
        self.homeui.buildingcombobox.clear()
        with open("buildings.txt", "r") as file_building:
            buildings = file_building.readlines()
        for building in buildings:
            self.homeui.buildingcombobox.addItem(building.strip())
        self.homeui.buildingcombobox.setCurrentIndex(index)

        index = self.homeui.buildingcombo_2.currentIndex()
        self.homeui.buildingcombo_2.clear()
        with open("buildings.txt", "r") as file_building:
            buildings = file_building.readlines()
        for building in buildings:
            self.homeui.buildingcombo_2.addItem(building.strip())

    def Add_Building_dialog(self, n):
        if n == "Show All Building":
            pass
        else:
            self.delete_widget(self.homeui.scrollAreaWidgetContents)
            for person_name, person_data in self.show_persons_by_building(n).items():
                self.person_layout = QWidget()
                self.person_widget = QGridLayout(self.person_layout)
                self.person_info = QLabel(f"Rental Name: {person_name}\nSerial No.:{person_data.get('Serial_Number')}\nBuilding:{person_data.get('Building')} ")
                self.Received_Rent = QLabel(f"Received Rent:{person_data.get('Received_Rent')}\nRent:{person_data.get('Rent')}\nBalance:{person_data.get('Balance_Rent')}")
                self.Electric_Bill = QPushButton(f"{person_data.get('Electric_Bill')}")
                self.Gas_Bill = QPushButton(f"{person_data.get('Gas_Bill')}")
                self.Gas_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                self.Electric_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                self.person_widget.addWidget(self.person_info, 0, 0, 9, 1)
                self.person_widget.addWidget(self.Received_Rent, 0, 1, 9, 1)
                self.person_widget.addWidget(self.Electric_Bill, 0, 2, 2, 1)
                self.person_widget.addWidget(self.Gas_Bill, 2, 2, 2, 1)
                self.homeui.scrollAreaWidgetContents.layout().addWidget(self.person_layout)
                if person_data.get("Electric_Bill") == "Electric Bill is not paid":
                    self.Gas_Bill.setStyleSheet(self.Gas_Bill.styleSheet() + "background-color: red;")
                elif person_data.get("Electric_Bill") == "Electric Bill is paid":
                    self.Electric_Bill.setStyleSheet(self.Electric_Bill.styleSheet() + "background-color: green;")

                if person_data.get("Gas_Bill") == "Gas Bill is not paid":
                    self.Gas_Bill.setStyleSheet(self.Gas_Bill.styleSheet() + "background-color: red;")
                elif person_data.get("Gas_Bill") == "Gas Bill is paid":
                    self.Electric_Bill.setStyleSheet(self.Electric_Bill.styleSheet() + "background-color: green;")

    def Add_Building(self, name: QLineEdit):
        with open("buildings.txt", "a") as file_building:
            file_building.write(name.text() + "\n")

    def setupui_search(self):
        if self.homeui.buildingcombobox.currentText() == "Show All Building":
            if self.homeui.searchedit.text() == "":
                self.delete_widget(self.homeui.scrollAreaWidgetContents)
                self.SetupUI()
        elif self.homeui.buildingcombobox.currentText() != "Show All Building":
            self.read_buildings_file()

    def search(self):
        search_text = self.homeui.searchedit.text()
        self.current_month, self.previous_month = Person.get_current_and_previous_month(self)
        try:
            with open(f"data/{self.current_month}/persondata.json", "r") as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            pass
        if search_text:
            for rentel_name, datas in data.items():
                if search_text.lower() in rentel_name.lower() or search_text.lower() in datas.get("Serial_Number") or search_text.lower() in datas.get("NIC") or search_text.lower() in datas.get("Due_Date"):
                    self.delete_widget(self.homeui.scrollAreaWidgetContents)
                    layout = self.homeui.scrollAreaWidgetContents.layout()
                    for key, value in datas.items():
                        label = QLabel(f"<h2>{key}:</h2><h3>{value}</h3>")
                        label.mousePressEvent = lambda _, p=datas: self.show_person_data(p)
                        label.setStyleSheet(label.styleSheet() + "*{background-color: #F2F3F3;}")
                        layout.addWidget(label)
                    self.SetupUI()
                    break
            else:
                QMessageBox.information(self, "Search Result", "No rental found with the given data")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def convert_to_excel(self, e):
        try:
            with open(f"data/{Person.get_current_and_previous_month(self)[0]}/persondata.json", "r") as json_file:
                data_dict = json.load(json_file)
            data_list = []
            for key, value in data_dict.items():
                value["Rentel_Name"] = key
                data_list.append(value)
            df = pd.DataFrame(data_list)
            excel_path = "RentData.xlsx"
            df.to_excel(excel_path, index=False)
            wb = load_workbook(excel_path)
            ws = wb.active
            header_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            header_font = Font(size=16, bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[cell.column_letter].width = 20
            wb.save(excel_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export data: {str(e)}")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def Add_Person_func(self):
        Person(
            self.homeui.Serial_Number.text(),
            self.homeui.NIC.text(),
            self.homeui.Rent.text(),
            self.homeui.Rentel_Name.text(),
            self.homeui.Due_Date.text(),
            self.homeui.Received_Rent.text(),
            self.homeui.Balance_Rent.text(),
            self.homeui.Electric_Bill.text(),
            self.homeui.Electricity_Meter_Number.text(),
            self.homeui.Electricity_Account_Number.text(),
            self.homeui.Consumer_Number.text(),
            self.homeui.Electricity_Meter_Name.text(),
            self.homeui.Gas_Costumer_Number.text(),
            self.homeui.Gas_Meter_Number.text(),
            self.homeui.Advance_Amount.text(),
            self.homeui.buildingcombo_2.currentText(),
            self.homeui.Gas_Bill.text(),
        )
        self.delete_widget(self.homeui.scrollAreaWidgetContents)
        self.SetupUI()
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def get_current_and_previous_month(self):
        today = datetime.date.today()
        current_month = today.strftime("%B_%Y")
        first_day_of_current_month = today.replace(day=1)
        last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(days=1)
        previous_month = last_day_of_previous_month.strftime("%B_%Y")
        return current_month, previous_month

    def SetupUI(self):
        if self.homeui.buildingcombobox.currentText() == "Show All Building":
            self.current_month, self.previous_month = self.get_current_and_previous_month()
            try:
                with open(f"data/{self.current_month}/persondata.json", "r") as json_file:
                    data = json.load(json_file)
            except FileNotFoundError:
                os.makedirs(f"data/{self.current_month}")
                with open(f"data/{self.current_month}/persondata.json", "w") as f:
                    dic = {"KK_Mosa": {"Plot_no": "72"}}
                    f.write(json.dumps(dic))
                with open(f"data/{self.current_month}/persondata.json", "r") as json_file:
                    data = json.load(json_file)
            for person_name, person_data in data.items():
                self.person_layout = QWidget()
                self.person_layout.setObjectName("person_layout")
                self.person_widget = QGridLayout(self.person_layout)
                self.person_info = QLabel(f"Rental Name: {person_name}\nSerial No.:{person_data.get('Serial_Number')}\nBuilding:{person_data.get('Building')} ")
                self.Received_Rent = QLabel(f"Received Rent:{person_data.get('Received_Rent')}\nRent:{person_data.get('Rent')}\nBalance:{person_data.get('Balance_Rent')}")
                self.Electric_Bill = QPushButton(f"{person_data.get('Electric_Bill')}")
                self.Gas_Bill = QPushButton(f"{person_data.get('Gas_Bill')}")
                self.person_layout.setStyleSheet("QWidget#person_layout{border-radius: 50px; padding: 1.5em; border: 3px dotted white;};")
                self.Gas_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                self.Electric_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                self.person_widget.addWidget(self.person_info, 0, 0, 9, 1)
                self.person_widget.addWidget(self.Received_Rent, 0, 1, 9, 1)
                self.person_widget.addWidget(self.Electric_Bill, 0, 2, 3, 1)
                self.person_widget.addWidget(self.Gas_Bill, 3, 2, 3, 1)
                self.homeui.scrollAreaWidgetContents.layout().addWidget(self.person_layout)
                if person_data.get("Due_Date") == str(datetime.datetime.now().day):
                    self.payment_info_widget = QWidget()
                    self.payment_info_widget.setStyleSheet("QWidget{border-radius: 50px; padding: 0.5em; background-color: rgba(255, 0, 0, 128);}; color: white;")
                    self.payment_info_layout = QVBoxLayout(self.payment_info_widget)
                    self.payment_info_label = QLabel(f"Payment Reminder: It's {person_data.get('Due_Date')}th and {person_name} didn't pay the rent.")
                    self.payment_info_layout.addWidget(self.payment_info_label)
                    self.homeui.scrollAreaWidgetContents.layout().addWidget(self.payment_info_widget)

    def electric_bill_method(self):
        if self.homeui.Electric_Bill.isChecked():
            self.homeui.Electric_Bill.setText("Electric Bill is paid")
            self.homeui.Electric_Bill.setStyleSheet(self.homeui.Electric_Bill.styleSheet() + "background-color: green;")
        else:
            self.homeui.Electric_Bill.setText("Electric Bill is not paid")
            self.homeui.Electric_Bill.setStyleSheet(self.homeui.Electric_Bill.styleSheet() + "background-color: red;")

    def gas_bill_method(self):
        if self.homeui.Gas_Bill.isChecked():
            self.homeui.Gas_Bill.setText("Gas Bill is paid")
            self.homeui.Gas_Bill.setStyleSheet(self.homeui.Gas_Bill.styleSheet() + "background-color: green;")
        else:
            self.homeui.Gas_Bill.setText("Gas Bill is not paid")
            self.homeui.Gas_Bill.setStyleSheet(self.homeui.Gas_Bill.styleSheet() + "background-color: red;")

    def electric_bill_3_method(self, form):
        if form.Electric_Bill_3.isChecked():
            form.Electric_Bill_3.setText("Electric Bill is paid")
            form.Electric_Bill_3.setStyleSheet(form.Electric_Bill_3.styleSheet() + "background-color: green;")
        else:
            form.Electric_Bill_3.setText("Electric Bill is not paid")
            form.Electric_Bill_3.setStyleSheet(form.Electric_Bill_3.styleSheet() + "background-color: red;")

    def gas_bill_3_method(self, form):
        if form.Gas_Bill_3.isChecked():
            form.Gas_Bill_3.setText("Gas Bill is paid")
            form.Gas_Bill_3.setStyleSheet(form.Gas_Bill_3.styleSheet() + "background-color: green;")
        else:
            form.Gas_Bill_3.setText("Gas Bill is not paid")
            form.Gas_Bill_3.setStyleSheet(form.Gas_Bill_3.styleSheet() + "background-color: red;")

    def taber(self, n):
        if n == 2:
            self.delete_widget(self.homeui.scrollAreaWidgetContents_expense)
            data = self.Expense_Sys.get_all_expenses()
            for expense in data:
                self.expense_title = QPushButton(f"{expense[0]}: RS.{expense[1]} Date Added:{expense[2]}")
                self.homeui.scrollAreaWidgetContents_expense.layout().addWidget(self.expense_title)

    def show_persons_by_building(self, building):
        try:
            with open(f"data/{self.get_current_and_previous_month()[0]}/persondata.json", "r") as file:
                data = json.load(file)
        except FileNotFoundError:
            data = {}
        try:
            filtered_persons = {name: info for name, info in data.items() if info["Building"] == building}
        except KeyError:
            filtered_persons = {}
        return filtered_persons

    def show_person_data(self, person):
        form = Ui_Form()
        widget = QDialog(self)
        form.setupUi(widget)
        if type(person) == dict:
            for key, value in person.items():
                person_widget = QLabel(f" <h2>{key}:</h2><h3>{value}</h3>")
                form.scrollAreaWidgetContents.layout().addWidget(person_widget)
            form.Serial_Number_3.setText(person["Serial_Number"])
            form.NIC_3.setText(person["NIC"])
            form.Rent_3.setText(person["Rent"])
            form.Rentel_Name_3.setText(person["Rentel_Name"])
            form.Due_Date_3.setText(person["Due_Date"])
            form.Received_Rent_3.setText(person["Received_Rent"])
            form.Balance_Rent_3.setText(person["Balance_Rent"])
            form.Electric_Bill_3.setText(person["Electric_Bill"])
            form.Electricity_Meter_Number_3.setText(person["Electricity_Meter_Number"])
            form.Electricity_Account_Number_3.setText(person["Electricity_Account_Number"])
            form.Consumer_Number_3.setText(person["Consumer_Number"])
            form.Electricity_Meter_Name_3.setText(person["Electricity_Meter_Name"])
            form.Gas_Costumer_Number_3.setText(person["Gas_Costumer_Number"])
            form.Gas_Meter_Number_3.setText(person["Gas_Meter_Number"])
            form.Advance_Amount_3.setText(person["Advance_Amount"])
            form.Gas_Bill_3.setText(person["Gas_Bill"])
            form.Building_3.setText(person["Building"])
            widget.setWindowTitle(person["Rentel_Name"])
            form.Gas_Bill_3.clicked.connect(lambda _, f=form: self.gas_bill_3_method(f))
            form.Electric_Bill_3.clicked.connect(lambda _, f=form: self.electric_bill_3_method(f))
            form.Add_Person_btn_3.clicked.connect(lambda _, f=form: self.update_person_in_json(f))
            widget.exec_()
        try:
            qt_material.apply_stylesheet(widget, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(widget, "light_teal.xml")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def update_person_in_json(self, form):
        person_data = {
            "Serial_Number": form.Serial_Number_3.text(),
            "NIC": form.NIC_3.text(),
            "Rent": form.Rent_3.text(),
            "Rentel_Name": form.Rentel_Name_3.text(),
            "Due_Date": form.Due_Date_3.text(),
            "Received_Rent": form.Received_Rent_3.text(),
            "Balance_Rent": form.Balance_Rent_3.text(),
            "Electric_Bill": form.Electric_Bill_3.text(),
            "Gas_Bill": form.Gas_Bill_3.text(),
            "Electricity_Meter_Number": form.Electricity_Meter_Number_3.text(),
            "Electricity_Account_Number": form.Electricity_Account_Number_3.text(),
            "Consumer_Number": form.Consumer_Number_3.text(),
            "Electricity_Meter_Name": form.Electricity_Meter_Name_3.text(),
            "Gas_Costumer_Number": form.Gas_Costumer_Number_3.text(),
            "Gas_Meter_Number": form.Gas_Meter_Number_3.text(),
            "Advance_Amount": form.Advance_Amount_3.text(),
            "Building": form.Building_3.text(),
        }
        try:
            with open(f"data/{self.get_current_and_previous_month()[0]}/persondata.json", "r") as file:
                data = json.load(file)
        except FileNotFoundError:
            data = []
        for person_name, person_datajson in data.items():
            if person_datajson["Serial_Number"] == person_data["Serial_Number"]:
                person_data["History"] = person_datajson["History"]
                data[person_name] = person_data
                break
        else:
            data.append(person_data)
        with open(f"data/{self.get_current_and_previous_month()[0]}/persondata.json", "w") as file:
            json.dump(data, file, indent=4)

    def closeEvent(self, event):
        self.setting.setValue("themeName", self.themeName)
        self.convert_to_excel(event)

    def remove_building(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Remove Building")
        line_edit = QLineEdit(dialog)
        line_edit.setPlaceholderText("Building Name To Remove...")
        line_edit.returnPressed.connect(lambda _="", n=line_edit.text(): self.remove_building_name(n))
        dialog.exec_()

    def remove_building_name(self, name):
        if not name:
            pass
        with open('buildings.txt', 'r') as file:
            lines = file.readlines()
        word_to_remove = name
        new_lines = []
        for line in lines:
            new_line = ' '.join([word for word in line.split() if word != word_to_remove])
            new_lines.append(new_line + "\n")
        with open('buildings.txt', 'w') as file:
            file.writelines(new_lines)

    def delete_widget(self, widget):
        for i in reversed(range(widget.layout().count())):
            widget_to_remove = widget.layout().itemAt(i).widget()
            if widget_to_remove is not None:
                widget_to_remove.setParent(None)

def main():
    reset_database()
    app = QApplication(sys.argv)
    window = XLSX()
    window.showMaximized()
    app.exec_()

main()

exit()
import os, sqlite3
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import qt_material, sys
from home import Ui_MainWindow
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from KK_Moosa_Plot_no_72 import Ui_Form

import sqlite3
from datetime import datetime as dt
import datetime
import json
import pandas as pd
class Database:
    def __init__(self, db_name):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.create_table()

    def create_table(self):
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS persons (
                id INTEGER PRIMARY KEY,
                serial_number TEXT,
                nic TEXT,
                rent TEXT,
                rentel_name TEXT,
                due_date TEXT,
                received_rent TEXT,
                balance_rent TEXT,
                electric_bill TEXT,
                gas_bill TEXT,
                electricity_meter_number TEXT,
                electricity_account_number TEXT,
                consumer_number TEXT,
                electricity_meter_name TEXT,
                gas_costumer_number TEXT,
                gas_meter_number TEXT,
                advance_amount TEXT,
                building TEXT,
                date_added TEXT
            )
        """
        )
        self.conn.commit()

    def add_person(
        self,
        serial_number,
        nic,
        rent,
        rentel_name,
        due_date,
        received_rent,
        balance_rent,
        electric_bill,
        gas_bill,
        electricity_meter_number,
        electricity_account_number,
        consumer_number,
        electricity_meter_name,
        gas_costumer_number,
        gas_meter_number,
        advance_amount,
        building,
    ):
        self.cursor.execute(
            """
            INSERT INTO persons (serial_number, nic, rent, rentel_name, due_date, received_rent, balance_rent, electric_bill, gas_bill, electricity_meter_number, electricity_account_number, consumer_number, electricity_meter_name, gas_costumer_number, gas_meter_number, advance_amount, building, date_added)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
            (
                serial_number,
                nic,
                rent,
                rentel_name,
                due_date,
                received_rent,
                balance_rent,
                electric_bill,
                gas_bill,
                electricity_meter_number,
                electricity_account_number,
                consumer_number,
                electricity_meter_name,
                gas_costumer_number,
                gas_meter_number,
                advance_amount,
                building,
                dt.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        self.conn.commit()

    def get_all_persons(self):
        self.cursor.execute("SELECT * FROM persons")
        return self.cursor.fetchall()

    def get_person_by_serial_number(self, serial_number):
        self.cursor.execute(
            "SELECT * FROM persons WHERE serial_number =?", (serial_number,)
        )
        return self.cursor.fetchone()

    def update_person(
        self,
        serial_number,
        nic,
        rent,
        rentel_name,
        due_date,
        received_rent,
        balance_rent,
        electric_bill,
        gas_bill,
        electricity_meter_number,
        electricity_account_number,
        consumer_number,
        electricity_meter_name,
        gas_costumer_number,
        gas_meter_number,
        advance_amount,
        building,
    ):
        self.cursor.execute(
            """
            UPDATE persons
            SET nic =?, rent =?, rentel_name =?, due_date =?, received_rent =?, balance_rent =?, electric_bill =?, gas_bill =?, electricity_meter_number =?, electricity_account_number =?, consumer_number =?, electricity_meter_name =?, gas_costumer_number =?, gas_meter_number =?, advance_amount =?, building =?
            WHERE serial_number =?
        """,
            (
                nic,
                rent,
                rentel_name,
                due_date,
                received_rent,
                balance_rent,
                electric_bill,
                gas_bill,
                electricity_meter_number,
                electricity_account_number,
                consumer_number,
                electricity_meter_name,
                gas_costumer_number,
                gas_meter_number,
                advance_amount,
                building,
                serial_number,
            ),
        )
        self.conn.commit()

    def close_connection(self):
        self.conn.close()


class Person:
    def __init__(
        self,
        Serial_Number,
        NIC,
        Rent,
        Rentel_Name,
        Due_Date,
        Received_Rent,
        Balance_Rent,
        Electric_Bill,
        Electricity_Meter_Number,
        Electricity_Account_Number,
        Consumer_Number,
        Electricity_Meter_Name,
        Gas_Costumer_Number,
        Gas_Meter_Number,
        Advance_Amount,
        Building,
        Gas_Bill,
    ):
        self.current_month, self.previous_month = self.get_current_and_previous_month(
            
        )
        self.setup_directories()
        today = datetime.date.today().strftime("%Y-%m-%d")
        self.db = Database("rent_data.db")
        self.db.add_person(
            Serial_Number,
            NIC,
            Rent,
            Rentel_Name,
            Due_Date,
            Received_Rent,
            Balance_Rent,
            Electric_Bill,
            Gas_Bill,
            Electricity_Meter_Number,
            Electricity_Account_Number,
            Consumer_Number,
            Electricity_Meter_Name,
            Gas_Costumer_Number,
            Gas_Meter_Number,
            Advance_Amount,
            Building,
        )

    def get_current_and_previous_month(self):
        today = datetime.date.today()
        current_month = today.strftime("%B_%Y")
        first_day_of_current_month = today.replace(day=1)
        last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(
            days=1
        )
        previous_month = last_day_of_previous_month.strftime("%B_%Y")
        return current_month, previous_month

    def setup_directories(self):
        if not os.path.exists("data"):
            os.makedirs("data")
        if not os.path.exists(f"data/{self.current_month}"):
            os.makedirs(f"data/{self.current_month}")
        if not os.path.exists(f"data/{self.previous_month}"):
            os.makedirs(f"data/{self.previous_month}")

    def add_expense(self, title, amount):
        self.db.cursor.execute(
            "INSERT INTO expenses (title, amount) VALUES (?,?)", (title, amount)
        )
        self.db.conn.commit()

    def get_all_expenses(self):
        self.db.cursor.execute("SELECT * FROM expenses")
        return self.db.cursor.fetchall()

    def get_expense_by_title(self, title):
        self.db.cursor.execute("SELECT * FROM expenses WHERE title =?", (title,))
        return self.db.cursor.fetchone()

    def update_expense(self, title, amount):
        self.db.cursor.execute(
            "UPDATE expenses SET amount =? WHERE title =?", (amount, title)
        )
        self.db.conn.commit()

    def close_connection(self):
        self.db.close_connection()


class XLSX(QMainWindow):
    def __init__(self):
        super().__init__()
        self.homeui = Ui_MainWindow()
        # self.Expense_Sys = Expenses()
        self.homeui.setupUi(self)
        self.homeui.Tab_window.currentChanged.connect(self.taber)
        self.homeui.action.triggered.connect(self.add_building)
        self.db = Database("rent_data.db")
        self.SetupUI()
        self.homeui.action_2.triggered.connect(self.remove_building)
        self.homeui.add_exp_btn.clicked.connect(self.add_exp_method)
        self.read_buildings_file()
        self.homeui.Electric_Bill.clicked.connect(self.electric_bill_method)
        self.homeui.Gas_Bill.clicked.connect(self.gas_bill_method)
        self.homeui.buildingcombobox.currentTextChanged.connect(
            self.Add_Building_dialog
        )
        self.homeui.menuPrint.triggered.connect(
            lambda: os.system("print " + f"RentData.xlsx")
        )
        self.setting = QSettings("Rent Recorder", "Theme")
        self.homeui.Add_Person_btn.clicked.connect(self.Add_Person_func)
        self.homeui.menuConvert_To_Excel.mousePressEvent = self.convert_to_excel
        self.homeui.searchedit.returnPressed.connect(self.search)
        self.themeName = ""
        self.searchtimer = QTimer(self.homeui.searchedit)
        self.searchtimer.timeout.connect(self.setupui_search)
        self.searchtimer.start(1000)
        self.homeui.dark_amber.changed.connect(self.Theme_Change)
        self.homeui.actiondark_blue.changed.connect(self.Theme_Change)
        self.homeui.actiondark_cyan.changed.connect(self.Theme_Change)
        self.homeui.actiondark_lightgreen.changed.connect(self.Theme_Change)
        self.homeui.actiondark_medical.changed.connect(self.Theme_Change)
        self.homeui.actiondark_pink.changed.connect(self.Theme_Change)
        self.homeui.actionlight_blue_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_cyan.changed.connect(self.Theme_Change)
        self.homeui.actionlight_cyan_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_lightgreen.changed.connect(self.Theme_Change)
        self.homeui.actionlight_lightgreen_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_orange.changed.connect(self.Theme_Change)
        self.homeui.actionlight_pink.changed.connect(self.Theme_Change)
        self.homeui.actionlight_pink_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_purple.changed.connect(self.Theme_Change)
        self.homeui.actionlight_purple_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_red.changed.connect(self.Theme_Change)
        self.homeui.actionlight_red_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_teal.changed.connect(self.Theme_Change)
        self.homeui.actionlight_teal_500.changed.connect(self.Theme_Change)
        self.homeui.actionlight_yellow.changed.connect(self.Theme_Change)
        self.homeui.actiondark_purple.changed.connect(self.Theme_Change)
        self.homeui.actiondark_red.changed.connect(self.Theme_Change)
        self.homeui.actiondark_teal.changed.connect(self.Theme_Change)
        self.homeui.actiondark_yellow.changed.connect(self.Theme_Change)
        self.homeui.actionlight_amber.changed.connect(self.Theme_Change)
        self.homeui.actionlight_blue.changed.connect(self.Theme_Change)
        self.homeui.actionlight_blue_500.changed.connect(self.Theme_Change)


    def Theme_Change(self):
        if self.homeui.dark_amber.isChecked():
            qt_material.apply_stylesheet(self, self.homeui.dark_amber.text() + ".xml")
            self.homeui.dark_amber.setChecked(False)
            self.themeName = self.homeui.dark_amber.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')



        elif self.homeui.actiondark_blue.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_blue.text() + ".xml"
            )
            self.homeui.actiondark_blue.setChecked(False)
            self.themeName = self.homeui.actiondark_blue.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_cyan.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_cyan.text() + ".xml"
            )
            self.homeui.actiondark_cyan.setChecked(False)
            self.themeName = self.homeui.actiondark_cyan.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_lightgreen.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_lightgreen.text() + ".xml"
            )
            self.homeui.actiondark_lightgreen.setChecked(False)
            self.themeName = self.homeui.actiondark_lightgreen.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_medical.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_medical.text() + ".xml"
            )
            self.homeui.actiondark_medical.setChecked(False)
            self.themeName = self.homeui.actiondark_medical.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_pink.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_pink.text() + ".xml"
            )
            self.homeui.actiondark_pink.setChecked(False)
            self.themeName = self.homeui.actiondark_pink.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_blue_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_blue_500.text() + ".xml"
            )
            self.homeui.actionlight_blue_500.setChecked(False)
            self.themeName = self.homeui.actionlight_blue_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_cyan.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_cyan.text() + ".xml"
            )
            self.homeui.actionlight_cyan.setChecked(False)
            self.themeName = self.homeui.actionlight_cyan.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_cyan_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_cyan_500.text() + ".xml"
            )
            self.homeui.actionlight_cyan_500.setChecked(False)
            self.themeName = self.homeui.actionlight_cyan_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_lightgreen.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_lightgreen.text() + ".xml"
            )
            self.homeui.actionlight_lightgreen.setChecked(False)
            self.themeName = self.homeui.actionlight_lightgreen.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_lightgreen_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_lightgreen_500.text() + ".xml"
            )
            self.homeui.actionlight_lightgreen_500.setChecked(False)
            self.themeName = self.homeui.actionlight_lightgreen_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_orange.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_orange.text() + ".xml"
            )
            self.homeui.actionlight_orange.setChecked(False)
            self.themeName = self.homeui.actionlight_orange.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_pink.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_pink.text() + ".xml"
            )
            self.homeui.actionlight_pink.setChecked(False)
            self.themeName = self.homeui.actionlight_pink.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_pink_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_pink_500.text() + ".xml"
            )
            self.homeui.actionlight_pink_500.setChecked(False)
            self.themeName = self.homeui.actionlight_pink_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_purple.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_purple.text() + ".xml"
            )
            self.homeui.actionlight_purple.setChecked(False)
            self.themeName = self.homeui.actionlight_purple.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_purple_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_purple_500.text() + ".xml"
            )
            self.homeui.actionlight_purple_500.setChecked(False)
            self.themeName = self.homeui.actionlight_purple_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_red.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_red.text() + ".xml"
            )
            self.homeui.actionlight_red.setChecked(False)
            self.themeName = self.homeui.actionlight_red.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_red_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_red_500.text() + ".xml"
            )
            self.homeui.actionlight_red_500.setChecked(False)
            self.themeName = self.homeui.actionlight_red_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_teal.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_teal.text() + ".xml"
            )
            self.homeui.actionlight_teal.setChecked(False)
            self.themeName = self.homeui.actionlight_teal.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_teal_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_teal_500.text() + ".xml"
            )
            self.homeui.actionlight_teal_500.setChecked(False)
            self.themeName = self.homeui.actionlight_teal_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_yellow.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_yellow.text() + ".xml"
            )
            self.homeui.actionlight_yellow.setChecked(False)
            self.themeName = self.homeui.actionlight_yellow.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_purple.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_purple.text() + ".xml"
            )
            self.homeui.actiondark_purple.setChecked(False)
            self.themeName = self.homeui.actiondark_purple.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_red.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_red.text() + ".xml"
            )
            self.homeui.actiondark_red.setChecked(False)
            self.themeName = self.homeui.actiondark_red.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_teal.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_teal.text() + ".xml"
            )
            self.homeui.actiondark_teal.setChecked(False)
            self.themeName = self.homeui.actiondark_teal.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actiondark_yellow.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actiondark_yellow.text() + ".xml"
            )
            self.homeui.actiondark_yellow.setChecked(False)
            self.themeName = self.homeui.actiondark_yellow.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_amber.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_amber.text() + ".xml"
            )
            self.homeui.actionlight_amber.setChecked(False)
            self.themeName = self.homeui.actionlight_amber.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_blue.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_blue.text() + ".xml"
            )
            self.homeui.actionlight_blue.setChecked(False)
            self.themeName = self.homeui.actionlight_blue.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        elif self.homeui.actionlight_blue_500.isChecked():
            qt_material.apply_stylesheet(
                self, self.homeui.actionlight_blue_500.text() + ".xml"
            )
            self.homeui.actionlight_blue_500.setChecked(False)
            self.themeName = self.homeui.actionlight_blue_500.text() + ".xml"
            self.setting.setValue("themeName", self.themeName)
            self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')


        
    def setupui_search(self):
        # print(self.homeui.searchedit.text())
        if self.homeui.buildingcombobox.currentText() == "Show All Building":
            if self.homeui.searchedit.text() == "":
                for i in reversed(range(self.homeui.scrollAreaWidgetContents.layout().count())):
                    widget_to_remove = self.homeui.scrollAreaWidgetContents.layout().itemAt(i).widget()
                    if widget_to_remove is not None:
                        widget_to_remove.setParent(None)

                self.SetupUI()
        elif self.homeui.buildingcombobox.currentText() != "Show All Building" :
            self.read_buildings_file()
        
        
        
    def search(self):
        search_text = self.homeui.searchedit.text()
        self.current_month, self.previous_month = Person.get_current_and_previous_month()
        
        try:
            with open(f"data/{self.current_month}/persondata.json", "r") as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            pass

        if search_text:
            for rentel_name, datas in data.items():
                if search_text.lower() in rentel_name.lower() or search_text.lower() in datas.get("Serial_Number") or search_text.lower() in datas.get("NIC") or search_text.lower() in datas.get("Due_Date"):
                    # QMessageBox.information(self, "Search",f" Search for {rentel_name},  in {datas}")
                    # self.show_person_data(datas)
                    for i in reversed(range(self.homeui.scrollAreaWidgetContents.layout().count())):
                        widget_to_remove = self.homeui.scrollAreaWidgetContents.layout().itemAt(i).widget()
                        if widget_to_remove is not None:
                            widget_to_remove.setParent(None)
                    layout = self.homeui.scrollAreaWidgetContents.layout()

                    
                    for key, value in datas.items():
                        label = QLabel(f"<h2>{key}:</h2><h3>{value}</h3>")
                        label.mousePressEvent = lambda _, p=datas: self.show_person_data(p)
                        label.setStyleSheet(label.styleSheet()+ "*{background-color: #F2F3F3;}")
                        layout.addWidget(label)
                        # layout.addWidget(QLabel("<hr>"))
                        

                    self.SetupUI()
                    break
            else:
                QMessageBox.information(self, "Search Result", "No rental found with the given data")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
            pass
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

        
    def convert_to_excel(self, e):
        try:
            with open(f"data/{Person.get_current_and_previous_month()[0]}/persondata.json", "r") as json_file:
                data_dict = json.load(json_file)
            data_list = []
            for key, value in data_dict.items():
                value["Rentel_Name"] = key
                data_list.append(value)

            df = pd.DataFrame(data_list)
            excel_path = "RentData.xlsx"
            df.to_excel(excel_path, index=False)

            # Load the workbook and access the sheet
            wb = load_workbook(excel_path)
            ws = wb.active

            # Apply style to the header row
            header_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            header_font = Font(size=16, bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[cell.column_letter].width = 20  # Adjust column width

            wb.save(excel_path)
            # QMessageBox.information(self, "Success", "Data successfully exported to RentData.xlsx")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export data: {str(e)}")
        try:
            qt_material.apply_stylesheet(self, self.setting.value("themeName"))
        except:
            qt_material.apply_stylesheet(self, "light_teal_500.xml")
            pass
        self.setStyleSheet(self.styleSheet() + '*{font: 11pt "Cascadia Code";}')

    def Add_Building_dialog(self, n):
        if n == "Show All Building":
            pass
            # self.setupui_search()pas
            # self.SetupUI()
        else:
            # print(type(widget_to_remove).__name__)

            for i in reversed(range(self.homeui.scrollAreaWidgetContents.layout().count())):
                widget_to_remove = self.homeui.scrollAreaWidgetContents.layout().itemAt(i).widget()
                print(type(widget_to_remove).__name__)
                if type(widget_to_remove).__name__ == "QLabel":
                    continue
                if widget_to_remove is not None:
                    widget_to_remove.setParent(None)
            for person_name, person_data in self.show_persons_by_building(n).items():
                # person_title = QPushButton(f"Rental Name: {person_name} Building: {person_data.get("Building")} Serial No.:{person_data.get("Serial_Number")}")
                # person_title.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                # # height
                # person_title.setStyleSheet(person_title.styleSheet()+ "height: 100px;")
                # self.homeui.scrollAreaWidgetContents.layout().addWidget(person_title)
                self.person_layout = QWidget()
                self.person_layout.setStyleSheet("QWidget{border-radius: 50px; padding: 0.5em; background-color: white;} QPushButton{border-radius:50px;}")
                self.person_widget = QGridLayout(self.person_layout)
                self.person_info = QLabel(f"Rental Name: {person_name}\nSerial No.:{person_data.get("Serial_Number")}\nBuilding:{person_data.get("Building")} ")
                self.Received_Rent = QLabel(f"Received Rent:{person_data.get("Received_Rent")}\nRent:{person_data.get("Rent")}\nBalance:{person_data.get("Balance_Rent")}")
                self.Electric_Bill = QPushButton(f"{person_data.get("Electric_Bill")}")
                self.Gas_Bill = QPushButton(f"{person_data.get("Gas_Bill")}")
                self.Gas_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                self.Electric_Bill.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                # self.Gas_Bill.setStyleSheet(sszelf.Gas_Bill.styleSheet()+ "border-radius: 50px;")
                # self.Electric_Bill.setStyleSheet(self.Electric_Bill.styleSheet()+ "border-radius: 50px;")
        

                # self.person_info.clicked.connect(lambda _, p=person_data: self.show_person_data(p))
                # height
                # self.person_info.setStyleSheet(self.person_info.styleSheet()+ "height: 100px;")
                self.person_widget.addWidget(self.person_info, 0, 0, 9,1)
                self.person_widget.addWidget(self.Received_Rent, 0, 1, 9,1)
                self.person_widget.addWidget(self.Electric_Bill, 0, 2, 2,1)
                self.person_widget.addWidget(self.Gas_Bill, 2, 2, 2,1)
                self.homeui.scrollAreaWidgetContents.layout().addWidget(self.person_layout)
                # self.person_layout.setStyleSheet(self.person_layout.styleSheet()+ "border: 1px solid red;")

                
                # for key, value in person_data.items():
                #     self.llabel = QLabel(f"<h1>{key}:</h1><h2>{value}</h2>")
                #     self.llabel.setHidden(True)
                #     # self.person_title.setText(f"<h1>{key}:</h1><h2>{value}</h2>")
                #     layout.addWidget(self.llabel)
                
                if person_data.get("Electric_Bill") == "Electric Bill is not paid":
                    self.Gas_Bill.setStyleSheet(self.Gas_Bill.styleSheet() + "background-color: red;")
                elif person_data.get("Electric_Bill") == "Electric Bill is paid":
                    self.Electric_Bill.setStyleSheet(self.Electric_Bill.styleSheet() + "background-color: green;")

                if person_data.get("Gas_Bill") =="Gas Bill is not paid":
                    self.Gas_Bill.setStyleSheet(self.Gas_Bill.styleSheet() + "background-color: red;")
                elif person_data.get("Gas_Bill") == "Gas Bill is paid":
                    self.Electric_Bill.setStyleSheet(self.Electric_Bill.styleSheet() + "background-color: green;")

            # print(self.show_persons_by_building(n))
            
        # Add building window
        # with open("buildings.txt", "a") as file_building:
        #     file_building.write()
        # pass


    def gas_bill_method(self):
        if self.homeui.Gas_Bill.isChecked():
            print("Yes", self.homeui.Gas_Bill.isChecked())
            self.homeui.Gas_Bill.setText("Gas Bill is paid")
            self.homeui.Gas_Bill.setStyleSheet(
                self.homeui.Gas_Bill.styleSheet() + "background-color: green;"
            )
        else:
            print("No", self.homeui.Gas_Bill.isChecked())
            self.homeui.Gas_Bill.setText("Gas Bill is not paid")
            self.homeui.Gas_Bill.setStyleSheet(
                self.homeui.Gas_Bill.styleSheet() + "background-color: red;"
            )

    def electric_bill_method(self):
        if self.homeui.Electric_Bill.isChecked():
            print("Yes", self.homeui.Electric_Bill.isChecked())
            self.homeui.Electric_Bill.setText("Electric Bill is paid")
            self.homeui.Electric_Bill.setStyleSheet(
                self.homeui.Electric_Bill.styleSheet() + "background-color: green;"
            )
        else:
            print("No", self.homeui.Electric_Bill.isChecked())
            self.homeui.Electric_Bill.setText("Electric Bill is not paid")
            self.homeui.Electric_Bill.setStyleSheet(
                self.homeui.Electric_Bill.styleSheet() + "background-color: red;"
            )

    def remove_building(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Remove Building")
        line_edit = QLineEdit(dialog)
        line_edit.setPlaceholderText("Building Name To Remove...")
        # dialog.layout().addWidget(line_edit)
        line_edit.returnPressed.connect(
            lambda _="", n=line_edit.text(): self.remove_building_name(n)
        )
        dialog.exec_()

    def add_building(self):
        dialog = QDialog(self)

        add_building_line_edit = QLineEdit(dialog)
        add_building_line_edit.returnPressed.connect(
            lambda _="", name=add_building_line_edit: self.Add_Building(name)
        )
        # dialog.layout().addWidget(add_building_line_edit)
        add_building_line_edit.setPlaceholderText("Enter Building Name")
        add_building_line_edit.move(10, 10)
        dialog.setWindowTitle("Add Building")
        add_building_line_edit.show()
        dialog.exec_()

    def read_buildings_file(self):
        index = self.homeui.buildingcombobox.currentIndex()
        self.homeui.buildingcombobox.clear()
        # self.homeui.buildingcombobox.addItem("Show All Building")
        # self.homeui.buildingcombobox.addItem("Add Building +")
        buildings = []
        with open("buildings.txt", "r") as file_building:
            buildings = file_building.readlines()
        for building in buildings:
            self.homeui.buildingcombobox.addItem(building.strip())
        self.homeui.buildingcombobox.setCurrentIndex(index)
        index = self.homeui.buildingcombo_2.currentIndex()

        self.homeui.buildingcombo_2.clear()

        buildings = []
        with open("buildings.txt", "r") as file_building:
            buildings = file_building.readlines()
        for building in buildings:
            self.homeui.buildingcombo_2.addItem(building.strip())
        # return buildings

    def taber(self, n):
        if n == 2:
            for i in reversed(
                range(self.homeui.scrollAreaWidgetContents_expense.layout().count())
            ):
                widget_to_remove = (
                    self.homeui.scrollAreaWidgetContents_expense.layout()
                    .itemAt(i)
                    .widget()
                )
                if widget_to_remove is not None:
                    widget_to_remove.setParent(None)
            data = self.Expense_Sys.get_all_expenses()
            for expense in data:
                self.expense_title = QPushButton(
                    f"{expense[0]}: RS.{expense[1]} Date Added:{expense[2]}"
                )
                self.homeui.scrollAreaWidgetContents_expense.layout().addWidget(
                    self.expense_title
                )

    def add_exp_method(self):
        title = self.homeui.title_lineedit.text()
        amount = self.homeui.exp_amount_lineedit.text()
        self.db.add_expense(title, int(amount))
        self.taber(2)

    def Add_Person_func(self):
        Person(
            self.homeui.Serial_Number.text(),
            self.homeui.NIC.text(),
            self.homeui.Rent.text(),
            self.homeui.Rentel_Name.text(),
            self.homeui.Due_Date.text(),
            self.homeui.Received_Rent.text(),
            self.homeui.Balance_Rent.text(),
            self.homeui.Electric_Bill.text(),
            self.homeui.Electricity_Meter_Number.text(),
            self.homeui.Electricity_Account_Number.text(),
            self.homeui.Consumer_Number.text(),
            self.homeui.Electricity_Meter_Name.text(),
            self.homeui.Gas_Costumer_Number.text(),
            self.homeui.Gas_Meter_Number.text(),
            self.homeui.Advance_Amount.text(),
            self.homeui.buildingcombo_2.currentText(),
            self.homeui.Gas_Bill.text(),
        )
        for i in reversed(range(self.homeui.scrollAreaWidgetContents.layout().count())):
            widget_to_remove = (
                self.homeui.scrollAreaWidgetContents.layout().itemAt(i).widget()
            )
            if widget_to_remove is not None:
                widget_to_remove.setParent(None)
        self.SetupUI()

    def show_persons_by_building(self, building):
        try:
            with open(f"data/{Person.get_current_and_previous_month()[0]}/persondata.json", "r") as file:
                data = json.load(file)
        except FileNotFoundError:
            data = {}
        try:
            filtered_persons = {name: info for name, info in data.items() if info["Building"] == building}
        except KeyError:
            print(f"No person found in building {building}")
            filtered_persons = {}

        return filtered_persons
        # self.display_persons(filtered_persons)


    def SetupUI(self):
        if self.homeui.buildingcombobox.currentText() == "Show All Building":
            data = self.db.get_all_persons()
            for person in data:
                self.person_layout = QWidget()
                self.person_layout.setStyleSheet(
                    "QWidget{border-radius: 50px; padding: 0.5em; background-color: white;} QPushButton{border-radius:50px;}"
                )
                self.person_widget = QGridLayout(self.person_layout)
                self.person_info = QLabel(
                    f"Rental Name: {person[3]}\nSerial No.:{person[0]}\nBuilding:{person[16]} "
                )
                self.Received_Rent = QLabel(
                    f"Received Rent:{person[5]}\nRent:{person[2]}\nBalance:{person[6]}"
                )
                self.Electric_Bill = QPushButton(f"{person[7]}")
                self.Gas_Bill = QPushButton(f"{person[8]}")
                self.Gas_Bill.clicked.connect(
                    lambda _, p=person: self.show_person_data(p)
                )
                self.Electric_Bill.clicked.connect(
                    lambda _, p=person: self.show_person_data(p)
                )
                self.person_widget.addWidget(self.person_info, 0, 0, 9, 1)
                self.person_widget.addWidget(self.Received_Rent, 0, 1, 9, 1)
                self.person_widget.addWidget(self.Electric_Bill, 0, 2, 2, 1)
                self.person_widget.addWidget(self.Gas_Bill, 2, 2, 2, 1)
                self.homeui.scrollAreaWidgetContents.layout().addWidget(
                    self.person_layout
                )
                if person[7] == "Electric Bill is not paid":
                    self.Electric_Bill.setStyleSheet(
                        self.Electric_Bill.styleSheet() + "background-color: red;"
                    )
                elif person[7] == "Electric Bill is paid":
                    self.Electric_Bill.setStyleSheet(
                        self.Electric_Bill.styleSheet() + "background-color: green;"
                    )
                if person[8] == "Gas Bill is not paid":
                    self.Gas_Bill.setStyleSheet(
                        self.Gas_Bill.styleSheet() + "background-color: red;"
                    )
                elif person[8] == "Gas Bill is paid":
                    self.Gas_Bill.setStyleSheet(
                        self.Gas_Bill.styleSheet() + "background-color: green;"
                    )

    def show_person_data(self, person):
        # print(type(person))
        form = Ui_Form()
        widget = QDialog(self)
        form.setupUi(widget)
        for key, value in zip(
            [
                "Serial Number",
                "NIC",
                "Rent",
                "Rentel Name",
                "Due Date",
                "Received Rent",
                "Balance Rent",
                "Electric Bill",
                "Gas Bill",
                "Electricity Meter Number",
                "Electricity Account Number",
                "Consumer Number",
                "Electricity Meter Name",
                "Gas Costumer Number",
                "Gas Meter Number",
                "Advance Amount",
                "Building",
            ],
            person,
        ):
            person_widget = QLabel(f" <h2>{key}:</h2><h3>{value}</h3>")
            person_widget.setStyleSheet(
                person_widget.styleSheet() + "*{background-color: #F2F3F3;}"
            )
            form.scrollAreaWidgetContents.layout().addWidget(person_widget)
        form.Serial_Number_3.setText(person[0])
        form.NIC_3.setText(person[1])
        form.Rent_3.setText(person[2])
        form.Rentel_Name_3.setText(person[3])
        form.Due_Date_3.setText(person[4])
        form.Received_Rent_3.setText(person[5])
        form.Balance_Rent_3.setText(person[6])
        form.Electric_Bill_3.setText(person[7])
        form.Gas_Bill_3.setText(person[8])
        form.Electricity_Meter_Number_3.setText(person[9])
        form.Electricity_Account_Number_3.setText(person[10])
        form.Consumer_Number_3.setText(person[11])
        form.Electricity_Meter_Name_3.setText(person[12])
        form.Gas_Costumer_Number_3.setText(person[13])
        form.Gas_Meter_Number_3.setText(person[14])
        form.Advance_Amount_3.setText(person[15])
        form.Building_3.setText(person[16])
        widget.setWindowTitle(person[3])
        form.Gas_Bill_3.clicked.connect(lambda _, f=form: self.gas_bill_3_method(f))
        form.Electric_Bill_3.clicked.connect(
            lambda _, f=form: self.electric_bill_3_method(f)
        )
        form.Add_Person_btn_3.clicked.connect(
            lambda _, f=form: self.update_person_in_json(f)
        )
        widget.exec_()

    def update_person_in_json(self, form):
        person_data = {
            "Serial Number": form.Serial_Number_3.text(),
            "NIC": form.NIC_3.text(),
            "Rent": form.Rent_3.text(),
            "Rentel Name": form.Rentel_Name_3.text(),
            "Due Date": form.Due_Date_3.text(),
            "Received Rent": form.Received_Rent_3.text(),
            "Balance Rent": form.Balance_Rent_3.text(),
            "Electric Bill": form.Electric_Bill_3.text(),
            "Gas Bill": form.Gas_Bill_3.text(),
            "Electricity Meter Number": form.Electricity_Meter_Number_3.text(),
            "Consumer Number": form.Consumer_Number_3.text(),
            "Electricity Meter Name": form.Electricity_Meter_Name_3.text(),
            "Gas Costumer Number": form.Gas_Costumer_Number_3.text(),
            "Gas Meter Number": form.Gas_Meter_Number_3.text(),
            "Advance Amount": form.Advance_Amount_3.text(),
            "Building": form.Building_3.text(),
        }

        self.db.cursor.execute(
            """
            UPDATE persons
            SET nic =?, rent =?, rentel_name =?, due_date =?, received_rent =?, balance_rent =?, electric_bill =?, gas_bill =?, electricity_meter_number =?, electricity_account_number =?, consumer_number =?, electricity_meter_name =?, gas_costumer_number =?, gas_meter_number =?, advance_amount =?, building =?
            WHERE serial_number =?
        """,
            (
                person_data["NIC"],
                person_data["Rent"],
                person_data["Rentel Name"],
                person_data["Due Date"],
                person_data["Received Rent"],
                person_data["Balance Rent"],
                person_data["Electric Bill"],
                person_data["Gas Bill"],
                person_data["Electricity Meter Number"],
                person_data["Electricity Account Number"],
                person_data["Consumer Number"],
                person_data["Electricity Meter Name"],
                person_data["Gas Costumer Number"],
                person_data["Gas Meter Number"],
                person_data["Advance Amount"],
                person_data["Building"],
                person_data["Serial Number"],
            ),
        )
        self.db.conn.commit()

    def electric_bill_3_method(self, form):
        if form.Electric_Bill_3.isChecked():
            print("Yes", form.Electric_Bill_3.isChecked())
            form.Electric_Bill_3.setText("Electric Bill is paid")
            form.Electric_Bill_3.setStyleSheet(
                form.Electric_Bill_3.styleSheet() + "background-color: green;"
            )
        else:
            print("No", form.Electric_Bill_3.isChecked())
            form.Electric_Bill_3.setText("Electric Bill is not paid")
            form.Electric_Bill_3.setStyleSheet(
                form.Electric_Bill_3.styleSheet() + "background-color: red;"
            )

    def gas_bill_3_method(self, form):
        if form.Gas_Bill_3.isChecked():
            print("Yes", form.Gas_Bill_3.isChecked())
            form.Gas_Bill_3.setText("Gas Bill is paid")
            form.Gas_Bill_3.setStyleSheet(
                form.Gas_Bill_3.styleSheet() + "background-color: green;"
            )
        else:
            print("No", form.Gas_Bill_3.isChecked())
            form.Gas_Bill_3.setText("Gas Bill is not paid")
            form.Gas_Bill_3.setStyleSheet(
                form.Gas_Bill_3.styleSheet() + "background-color: red;"
            )


def main():
    app = QApplication(sys.argv)
    window = XLSX()
    # setupUI(window)
    window.showMaximized()
    app.exec_()


main()
