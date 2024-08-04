import pyautogui as gui
import time
from PIL import Image
from pytesseract import pytesseract
import numpy as np

# import matplotlib.pyplot as plt

# import matplotlib.image as mpimg


def rgb2gray(rgb):
    return np.dot(rgb[..., :3], [0.2989, 0.5870, 0.1140])


def has_numbers(input_string):
    digits = "".join(char for char in input_string if char.isdigit())
    return bool(digits), digits


is_on_right = False
is_on_left = False

while True:
    gui.keyDown("w")
    # gui.keyDown("a")
    # is_on_right = False
    # is_on_left = False
    time.sleep(3)
    # Defining paths to tesseract.exe
    # and the image we would be using
    path_to_tesseract = r"C:\Users\Black\AppData\Local\Tesseract-OCR\tesseract.exe"
    gui.screenshot("image.png", (1120, 620, 100, 50))
    image_path = "image.png"
    # Opening the image & storing it in an image object
    img = Image.open(image_path)
    pytesseract.tesseract_cmd = path_to_tesseract

    img_array = np.array(img)
    gray = rgb2gray(img_array)
    gray_image = Image.fromarray(np.uint8(gray))
    # Passing the image object to image_to_string() function
    # This function will extract the text from the image
    text = pytesseract.image_to_string(gray_image)
    # Example usag
    print(text[:-1])
    if text[:-1]:
        contains_number, combined_digits = has_numbers(f"{text[:-1]}")
        print(contains_number)  # True
        print((combined_digits))  # 12
        try:
            if int(combined_digits) <= 100:
                # print("YES>>>Yes>>Yes>>Yes>>Yes>>Yes")
                if is_on_right == False:
                    gui.keyDown("d")
                    print("D")
                    gui.moveTo(1255, 255)
                    is_on_right = True
                    is_on_left = False
                    # time.sleep(5)
                if is_on_left == False:
                    gui.keyDown("a")
                    gui.moveTo(255, 255)
                    print("A")
                    is_on_right = False
                    is_on_left = True
                    # time.sleep(5)
        except:
            pass

exit()
import requests
from lxml import html


def get_fba_sellers(asin):
    url = f"https://www.amazon.com/dp/B000VZJ028/ref=olp-opf-redir?aod=1"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    tree = html.fromstring(response.content)
    sellers = tree.xpath(
        '//div[contains(@class, "olpOffer")]//h3[contains(@class, "olpSellerName")]/span/a/text()'
    )
    return sellers


asin = "B000VZJ028"  # Example ASIN
sellers = get_fba_sellers(asin)
print(sellers)

exit()

import sys, datetime
import os
import json
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QMessageBox,
    QListWidget,
    QListWidgetItem,
    QFileDialog,
)
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *


import sys
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QPushButton,
    QDialog,
    QLabel,
    QHBoxLayout,
    QFormLayout,
)


class PersonWidget(QWidget):
    def __init__(self, name, date, amount):
        super().__init__()
        self.name = name
        self.date = date
        self.amount = amount
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout()
        layout.addRow("Name:", QLabel(self.name))
        layout.addRow("Date:", QLabel(self.date))
        layout.addRow("Amount:", QLabel(str(self.amount)))
        self.setLayout(layout)
        self.setWindowTitle(self.name)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Person Data Widget")
        self.setGeometry(100, 100, 400, 300)

        self.init_ui()

    def init_ui(self):
        # Main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)

        # Example data - list of persons
        persons = [
            {"name": "Alice", "date": "2023-01-15", "amount": 1000},
            {"name": "Bob", "date": "2023-02-20", "amount": 1500},
            {"name": "Charlie", "date": "2023-03-25", "amount": 1200},
        ]

        # Create buttons for each person
        for person in persons:
            button = QPushButton(person["name"])
            button.clicked.connect(lambda _, p=person: self.show_person_data(p))
            main_layout.addWidget(button)

    def show_person_data(self, person):
        dialog = QDialog(self)
        person_widget = PersonWidget(person["name"], person["date"], person["amount"])
        dialog.layout = QVBoxLayout()
        dialog.layout.addWidget(person_widget)
        dialog.setLayout(dialog.layout)
        dialog.setWindowTitle(person["name"])
        dialog.exec_()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

exit()
from openpyxl import workbook, load_workbook
import os
from openpyxl.styles import Font, Alignment

wb = load_workbook("Rent sheet  JUNE(AutoRecovered).xlsx")
ws = wb.active


def change_color(color):
    # print all Names that is under A
    for cell in ws["A"]:
        if cell.value is not None:
            # print(cell.value)
            # Change the cell color to red
            cell.font = Font(color=color, bold=True, size=24)
            # change the cell size to fit the cell's contents
            # do this here with cell


def change_style():
    # Define the new font size
    new_font_size = 24
    new_font = Font(size=new_font_size, bold=True)
    aling = Alignment("center", "center")
    # Iterate through all rows and columns to change the font size
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # Check if the cell is not empty
                cell.font = new_font
                ws.row_dimensions[cell.row].height = 50
                ws.column_dimensions[cell.column_letter].width = 50
                cell.alignment = aling


def getotp():
    # Get the maximum row number with data
    max_row = ws.max_row
    rows_data = []
    # Iterate through the rows and print data from row 3 to max_row
    for row in ws.iter_rows(min_row=3, max_row=max_row, max_col=3, values_only=True):
        rows_data.append(row)

    return rows_data


otp_data = getotp()


def get_rent_by_name(rental_name: str):
    """_summary_

    Args:
        rental_name (str): rental name here
    """

    # Iterate through the rows to find the rental name
    # max_row = ws.max_row

    for row in otp_data:
        print(row)
        if row[2] == rental_name:
            rent = row[1]
            print(f"The rent for {rental_name} is {rent}")
            return
    print(f"Rental name '{rental_name}' not found.")


def show_rental_data(otp_data: list):
    """_summary_

    Args:
        otp_data (list): otp data shows output like this (1, 12000, 'Tailor')
    """

    # Ask the user for the rental name
    for i, row in enumerate(otp_data):
        if row != (None, None, None):
            if row != (None, "=SUM(B3:B10)", None):
                if row != ("Balance rent of the month", None, None):
                    print(i, row)


def main():
    change_color("ff000000")
    change_style()
    otp_data = getotp()
    show_rental_data(otp_data)

    User_input = input("Enter the rental name: ")
    get_rent_by_name(User_input)

    # os.system("cls")
    # print B11 collumn
    print(ws["B11"].value)

    # ws.merge_cells("B12:C12")
    ws.delete_cols(3, 1)

    # Add formula to cell C12 column number 3 Sum(B11:E10):Sum of the cells
    ws.cell(12, 3).value = "=Sum(B11-D11)"

    # Get
    # now save it
    wb.save("Rent sheet  JUNE(AutoRecovered) - Updated.xlsx")


if __name__ == "__main__":
    main()
